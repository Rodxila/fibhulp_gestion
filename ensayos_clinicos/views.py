from django.shortcuts import render, redirect
from .models import DatosEnsayo, FacturaMensual
from .forms import DatosEnsayoForm
from django.contrib import messages
import json
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import openpyxl
from openpyxl import load_workbook
from django.http import HttpResponse
from datetime import datetime
from collections import Counter
import re
from .forms import LoginForm
from .models import Usuario
from collections import defaultdict
from .models import FacturaDetalle
from rrhh.models import FormularioContratacion
from functools import wraps

def restringir_seccion(seccion_requerida):
    def decorador(vista_func):
        @wraps(vista_func)
        def _wrapped_view(request, *args, **kwargs):
            usuario_id = request.session.get('usuario_id')
            if not usuario_id:
                return redirect('login')

            from ensayos_clinicos.models import Usuario
            try:
                usuario = Usuario.objects.get(id=usuario_id)
                if str(seccion_requerida) not in usuario.secciones.split(','):
                    messages.error(request, "No tienes acceso a esta sección.")
                    return redirect('login')
            except Usuario.DoesNotExist:
                return redirect('login')

            return vista_func(request, *args, **kwargs)
        return _wrapped_view
    return decorador




@restringir_seccion('1')
def contratos_view(request):
    if not request.session.get('usuario_id'):
        return redirect('login')
    meses = [mes for mes, _ in DatosEnsayo._meta.get_field('mes').choices]
    años_disponibles = DatosEnsayo.objects.values_list('año', flat=True).distinct().order_by('año')
    if request.method == 'POST':
        form = DatosEnsayoForm(request.POST)
        mes = request.POST.get('mes')
        año = request.POST.get('año')

        if 'delete-month' in request.POST:
            if mes and año:
                if DatosEnsayo.objects.filter(mes=mes, año=año).exists():
                    DatosEnsayo.objects.filter(mes=mes, año=año).delete()
                    messages.success(request, f"Se han eliminado los datos de {mes} {año}.")
                else:
                    messages.warning(request, f"No se encontraron datos para {mes} {año}.")
            else:
                messages.error(request, "Debe indicar un mes y un año para borrar.")
            return redirect('contratos')

        if form.is_valid():
            obj, creado = DatosEnsayo.objects.update_or_create(
                mes=form.cleaned_data['mes'],
                año=form.cleaned_data['año'],
                defaults={
                    'ensayos': form.cleaned_data['ensayos'],
                    'observacionales': form.cleaned_data['observacionales'],
                    'adendas': form.cleaned_data['adendas']
                }
            )
            messages.success(request, f"Registro de {form.cleaned_data['mes']} {form.cleaned_data['año']} {'creado' if creado else 'actualizado'}.")
            return redirect('contratos')

    else:
        form = DatosEnsayoForm()

    datos = DatosEnsayo.objects.all().order_by('año', 'mes')
    chart_data = {
        'labels': [f"{d.mes} {d.año}" for d in datos],
        'ensayos': [d.ensayos for d in datos],
        'observacionales': [d.observacionales for d in datos],
        'adendas': [d.adendas for d in datos],
    }

    total_ensayos = sum(d.ensayos for d in datos)
    total_obs = sum(d.observacionales for d in datos)
    total_adendas = sum(d.adendas for d in datos)

    # Comparación
    ano_1 = request.GET.get('comparar_ano_1')
    ano_2 = request.GET.get('comparar_ano_2')
    comparacion = None

    if ano_1 and ano_2 and ano_1 != ano_2:
        try:
            ano_1 = int(ano_1)
            ano_2 = int(ano_2)
            datos_1 = {r.mes: r for r in DatosEnsayo.objects.filter(año=ano_1)}
            datos_2 = {r.mes: r for r in DatosEnsayo.objects.filter(año=ano_2)}

            comparacion = {
                'año_1': ano_1,
                'año_2': ano_2,
                'labels': meses,
                'ensayos_1': [datos_1.get(m).ensayos if datos_1.get(m) else 0 for m in meses],
                'ensayos_2': [datos_2.get(m).ensayos if datos_2.get(m) else 0 for m in meses],
                'obs_1': [datos_1.get(m).observacionales if datos_1.get(m) else 0 for m in meses],
                'obs_2': [datos_2.get(m).observacionales if datos_2.get(m) else 0 for m in meses],
                'adendas_1': [datos_1.get(m).adendas if datos_1.get(m) else 0 for m in meses],
                'adendas_2': [datos_2.get(m).adendas if datos_2.get(m) else 0 for m in meses],
            }
        except ValueError:
            comparacion = None

    # Estructuras para totales por mes y año
    estructura = defaultdict(lambda: defaultdict(lambda: {'ensayos': 0, 'observacionales': 0, 'adendas': 0}))
    totales = defaultdict(lambda: {'ensayos': 0, 'observacionales': 0, 'adendas': 0})

    for d in datos:
        estructura[d.mes][d.año] = {
            'ensayos': d.ensayos,
            'observacionales': d.observacionales,
            'adendas': d.adendas
        }
        totales[d.año]['ensayos'] += d.ensayos
        totales[d.año]['observacionales'] += d.observacionales
        totales[d.año]['adendas'] += d.adendas

    return render(request, 'ensayos_clinicos/contratos.html', {
    'form': form,
    'chart_data': json.dumps(chart_data),
    'total_ensayos': total_ensayos,
    'total_obs': total_obs,
    'total_adendas': total_adendas,
    'comparacion': comparacion,
    'años_disponibles': años_disponibles,
    'datos': estructura,
    'totales': totales,

})



def bienvenida_view(request):
    return render(request, 'ensayos_clinicos/bienvenida.html')

@restringir_seccion('1')
def facturas_view(request):
    if not request.session.get('usuario_id'):
        return redirect('login')

    from .models import FacturaDetalle

    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    comparacion = None

    if request.method == 'POST':
        if 'subir_excel' in request.POST and request.FILES.get('archivo_excel'):
            archivo = request.FILES['archivo_excel']
            wb = load_workbook(archivo, data_only=True)
            ws = wb.worksheets[0]
            mensual = {}
            contador = 0
            mes_actual = None
            año_actual = int(request.POST.get('año') or 0)

            # Borrado previo
            FacturaMensual.objects.filter(año=año_actual).delete()
            FacturaDetalle.objects.filter(año=año_actual).delete()

            regex_total = re.compile(r'total\s+(\w+)', re.IGNORECASE)
            meses_normalizados = {
                'enero': 'Ene', 'febrero': 'Feb', 'marzo': 'Mar', 'abril': 'Abr',
                'mayo': 'May', 'junio': 'Jun', 'julio': 'Jul', 'agosto': 'Ago',
                'septiembre': 'Sep', 'octubre': 'Oct', 'noviembre': 'Nov', 'diciembre': 'Dic'
            }

            for row in ws.iter_rows(min_row=5):
                celda_a = str(row[0].value).strip().lower() if row[0].value else ""
                celda_base_total = row[4].value
                celda_fecha_cobro = row[6].value
                fecha_cobro = celda_fecha_cobro.date() if isinstance(celda_fecha_cobro, datetime) else None

                match = regex_total.search(celda_a)
                if match:
                    nombre_mes = match.group(1).lower()
                    mes_actual = meses_normalizados.get(nombre_mes)
                    mensual[mes_actual] = {
                        'cantidad': float(celda_base_total or 0),
                        'n': contador
                    }
                    contador = 0
                else:
                    contador += 1
                    if mes_actual:
                        FacturaDetalle.objects.create(
                            año=año_actual,
                            mes=mes_actual,
                            identificador=str(row[2].value).strip() if row[2].value else "",
                            concepto=str(row[3].value).strip() if row[3].value else "",
                            cantidad=celda_base_total or 0,
                            fecha_emision=row[1].value if isinstance(row[1].value, datetime) else None,
                            fecha_ultimo_cobro=fecha_cobro
                        )

            total = 0
            for mes, datos in mensual.items():
                FacturaMensual.objects.create(
                    año=año_actual,
                    mes=mes,
                    cantidad=datos['cantidad'],
                    numero_facturas=datos['n'],
                    fecha_ultimo_cobro=fecha_cobro
                )
                total += datos['cantidad']

            messages.success(request, f"Se ha importado correctamente el Excel para {año_actual}.")
            return redirect('facturas')

        if 'delete-year' in request.POST:
            año = int(request.POST.get('año') or 0)
            if año:
                eliminados = FacturaMensual.objects.filter(año=año).count()
                if eliminados:
                    FacturaMensual.objects.filter(año=año).delete()
                    messages.success(request, f"Se eliminaron {eliminados} registros del año {año}.")
                else:
                    messages.warning(request, f"No hay registros para el año {año}.")
            else:
                messages.error(request, "Año no válido.")
            return redirect('facturas')

        # Procesamiento manual desde el formulario
        meta = int(request.POST.get('meta') or 0)
        año = int(request.POST.get('año') or 0)
        mensual = {}
        total = 0

        FacturaMensual.objects.filter(año=año).delete()
        for mes in meses:
            valor = float(request.POST.get(mes, 0) or 0)
            n_facturas = int(request.POST.get(f"n_{mes}", 0) or 0)
            mensual[mes] = {'cantidad': valor, 'n': n_facturas}
            total += valor
            FacturaMensual.objects.create(año=año, mes=mes, cantidad=valor, numero_facturas=n_facturas)

    # Bloques de facturación
    registros = FacturaMensual.objects.all().order_by('año', 'mes')
    años_disponibles = sorted(set(registros.values_list('año', flat=True)))
    bloques_facturas = []

    for año in años_disponibles:
        mensual = {}
        total = 0
        for r in registros.filter(año=año).values('mes', 'cantidad', 'numero_facturas'):
            try:
                cantidad = Decimal(str(r['cantidad'])).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                n = int(r['numero_facturas'] or 0)
                mensual[r['mes']] = {'cantidad': float(cantidad), 'n': n}
                total += float(cantidad)
            except:
                continue

        medias = [
            round(mensual.get(m, {'cantidad': 0, 'n': 1})['cantidad'] / mensual.get(m, {'cantidad': 0, 'n': 1})['n'], 2)
            if mensual.get(m, {'n': 0})['n'] else 0 for m in meses
        ]

        chart_data = {
            'labels': [f"{m} {año}" for m in meses],
            'valores': [mensual.get(m, {'cantidad': 0})['cantidad'] for m in meses],
            'medias': medias
        }

        bloques_facturas.append({
            'año': año,
            'mensual': mensual,
            'total': total,
            'progreso': 0,
            'medias': medias,
            'chart_data': chart_data
        })

    # Comparación de años
    ano_1 = request.GET.get('comparar_ano_1')
    ano_2 = request.GET.get('comparar_ano_2')
    if ano_1 and ano_2:
        try:
            ano_1 = int(ano_1)
            ano_2 = int(ano_2)
            datos_1 = {r.mes: r for r in FacturaMensual.objects.filter(año=ano_1)}
            datos_2 = {r.mes: r for r in FacturaMensual.objects.filter(año=ano_2)}

            comparacion = {
                'año_1': ano_1,
                'año_2': ano_2,
                'labels': meses,
                'valores_1': [float(datos_1.get(m).cantidad) if datos_1.get(m) else 0 for m in meses],
                'valores_2': [float(datos_2.get(m).cantidad) if datos_2.get(m) else 0 for m in meses],
            }
        except:
            comparacion = None

    context = {
        'meta': 0,
        'meses': meses,
        'bloques_facturas': bloques_facturas,
        'bloques_facturas_json': json.dumps([
            {'año': b['año'], 'chart_data': b['chart_data']} for b in bloques_facturas
        ]),
        'comparacion': comparacion,
        'años_disponibles': años_disponibles,
        'facturas_detalle': FacturaDetalle.objects.all().order_by('-año', 'mes')
    }

    return render(request, 'ensayos_clinicos/facturas.html', context)

def exportar_facturas_excel(request):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    años = FacturaMensual.objects.values_list('año', flat=True).distinct().order_by('año')

    for año in años:
        ws = wb.create_sheet(title=f"Facturas_{año}")
        ws.append(["Mes", "Cantidad (€)", "Nº Facturas", "Media (€)"])

        registros = FacturaMensual.objects.filter(año=año).order_by('mes')
        for r in registros:
            media = r.cantidad / r.numero_facturas if r.numero_facturas else 0
            ws.append([r.mes, formato_europeo(r.cantidad), r.numero_facturas, formato_europeo(media)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Facturas_FIBHULP.xlsx'
    wb.save(response)
    return response


def exportar_contratos_excel(request):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    años = DatosEnsayo.objects.values_list('año', flat=True).distinct().order_by('año')
    meses = [mes for mes, _ in DatosEnsayo._meta.get_field('mes').choices]

    for año in años:
        ws = wb.create_sheet(title=f"Contratos_{año}")
        ws.append(["Mes", "Ensayos", "Observacionales", "Adendas"])

        registros = DatosEnsayo.objects.filter(año=año)
        for mes in meses:
            r = registros.filter(mes=mes).first()
            if r:
                ws.append([mes, r.ensayos, r.observacionales, r.adendas])

            else:
                ws.append([mes, 0, 0, 0])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Contratos_FIBHULP.xlsx'
    wb.save(response)
    return response


def formato_europeo(valor):
    try:
        valor = float(valor)
        return "{:,.2f}".format(valor).replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return valor

from django.contrib.auth.hashers import check_password
from rrhh.models import FormularioContratacion

def is_password_hashed(pwd):
    return pwd.startswith('pbkdf2_') or pwd.startswith('bcrypt_') or pwd.startswith('argon2')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        try:
            user = Usuario.objects.get(username=username)

            # Verificación segura: hasheado o texto plano
            if is_password_hashed(user.password):
                if not check_password(password, user.password):
                    raise Usuario.DoesNotExist
            else:
                if user.password != password:
                    raise Usuario.DoesNotExist

            # Guardar sesión con claves unificadas
            request.session['usuario_id'] = user.id
            request.session['usuario_nombre'] = user.username
            request.session['secciones'] = user.secciones

            # Redirección según permisos
            if '1' in user.secciones:
                return redirect('facturas')
            elif '2' in user.secciones:
                return redirect('legal_concursos')
            elif '3' in user.secciones:
                return redirect('rrhh:rrhh_inicio')
            elif '4' in user.secciones:
                form = FormularioContratacion.objects.filter(empleado=user).first()
                if form:
                    return redirect('rrhh:formulario_empleado', pk=form.pk)
                else:
                    messages.error(request, "No se encontró ningún formulario asignado.")
                    return redirect('login')
            else:
                messages.error(request, "No tienes acceso a ninguna sección.")
                return redirect('login')

        except Usuario.DoesNotExist:
            messages.error(request, "Usuario o contraseña incorrectos.")

    return render(request, 'ensayos_clinicos/login.html')


def logout_view(request):
    request.session.flush()
    messages.info(request, 'Sesión cerrada correctamente')
    return redirect('login')



@restringir_seccion('1')
def facturas_detalle_view(request):
    facturas = FacturaDetalle.objects.all()

    # Filtros
    año = request.GET.get('año')
    mes = request.GET.get('mes')
    identificador = request.GET.get('identificador')

    if año:
        facturas = facturas.filter(año=año)
    if mes:
        facturas = facturas.filter(mes__iexact=mes)
    if identificador:
        facturas = facturas.filter(identificador__icontains=identificador)

    años = FacturaDetalle.objects.values_list('año', flat=True).distinct().order_by('año')
    meses = FacturaDetalle.objects.values_list('mes', flat=True).distinct().order_by('mes')

    return render(request, 'ensayos_clinicos/facturas_detalle.html', {
        'facturas': facturas,
        'años': años,
        'meses': meses
    })

