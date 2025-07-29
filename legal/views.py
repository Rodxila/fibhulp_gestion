import pandas as pd
from django.shortcuts import render, redirect
from .models import Convocatoria, Concurso
from .forms import ConvocatoriaForm
from collections import defaultdict
from datetime import datetime, timedelta, date
import re
from django.utils.timezone import now
from ensayos_clinicos.models import Usuario
from functools import wraps
from django.contrib import messages
from .forms import ConcursoForm
from django.core.mail import send_mail
from .utils import extraer_fecha

def enviar_correos_concursos():
    from .models import Concurso
    concursos = Concurso.objects.all()

    for c in concursos:
        datos = evaluar_concurso_para_correo(c)
        if datos:
            send_mail(
                datos['asunto'],
                datos['mensaje'],
                'notificacionesfib@gmail.com',
                [datos['correo']],
                fail_silently=False,
            )



def evaluar_concurso_para_correo(c):
    estado = (c.ESTADO or '').strip().upper()
    print(f"📌 Evaluando: {c.EXPEDIENTE} | Estado: {estado}")
    if estado in ['CANCELADO', 'FINALIZADO', 'PARALIZADO', 'DESIERTO Y FINALIZADO', 'EN GESTIÓN']:
        print("❌ Estado excluido.")
        return None

    hoy = date.today()
    prorroga = (c.PRORROGA or '').lower()
    fecha_fin = extraer_fecha(c.FECHA_FIN)
    fecha_prorroga = extraer_fecha(c.FECHA_FIN_PRORROGA)
    dias_restantes = dias_prorroga = dias_intermedio = None
    asunto = mensaje = frecuencia = None

    print(f"🔎 Prórroga: {prorroga}, Fecha fin: {fecha_fin}, Fecha prórroga: {fecha_prorroga}")

    if 'si' in prorroga:
        if fecha_fin:
            dias_restantes = (fecha_fin - hoy).days
            print(f"📅 Días hasta fin (para firmar prórroga): {dias_restantes}")
            if 60 < dias_restantes <= 90:
                asunto = "📅 Fin de prórroga cercano"
                mensaje = f"El concurso '{c.EXPEDIENTE}' tiene una prórroga que debe firmarse y finaliza en {dias_restantes} días."
                frecuencia = "semanal"
            elif 7 < dias_restantes <= 60:
                asunto = "✍️ Se requiere firma de contrato por prórroga"
                mensaje = f"Debe firmarse la prórroga del contrato '{c.EXPEDIENTE}'. Quedan {dias_restantes} días."
                frecuencia = "semanal"
            elif 0 < dias_restantes <= 7:
                asunto = "⚠️ URGENTE: Prórroga debe firmarse ya"
                mensaje = f"La prórroga del concurso '{c.EXPEDIENTE}' debe firmarse inmediatamente. Menos de una semana."
                frecuencia = "diaria"

        if fecha_fin and fecha_prorroga:
            diferencia = (fecha_prorroga - fecha_fin).days
            intermedia = fecha_fin + timedelta(days=round(diferencia / 2))
            tolerancia = 7
            dias_intermedio = (intermedia - hoy).days
            print(f"🟰 Punto medio entre fechas: {intermedia}, Días al punto medio: {dias_intermedio}")
            if abs(dias_intermedio) <= tolerancia:
                asunto = "🔁 Renovación intermedia de prórroga"
                mensaje = f"Se cumple un año desde la primera prórroga del concurso '{c.EXPEDIENTE}'. Evaluar renovación."
                frecuencia = "semanal"

        if fecha_prorroga:
            dias_prorroga = (fecha_prorroga - hoy).days
            print(f"📅 Días hasta fin de prórroga: {dias_prorroga}")
            if 60 < dias_prorroga <= 90:
                asunto = "📅 Fin de contrato prorrogado cercano"
                mensaje = f"El concurso '{c.EXPEDIENTE}' prorrogado finaliza en {dias_prorroga} días."
                frecuencia = "semanal"
            elif 7 < dias_prorroga <= 60:
                asunto = "✍️ Revisión final de contrato prorrogado"
                mensaje = f"Revisar el cierre del contrato '{c.EXPEDIENTE}' prorrogado. Quedan {dias_prorroga} días."
                frecuencia = "semanal"
            elif 0 < dias_prorroga <= 7:
                asunto = "⚠️ URGENTE: Contrato prorrogado finaliza ya"
                mensaje = f"El contrato prorrogado del concurso '{c.EXPEDIENTE}' finaliza esta semana."
                frecuencia = "diaria"

    else:
        if fecha_fin:
            dias_restantes = (fecha_fin - hoy).days
            print(f"📅 Días hasta fin sin prórroga: {dias_restantes}")
            if 60 < dias_restantes <= 120:
                asunto = "📝 Iniciar redacción del pliego"
                mensaje = f"El contrato '{c.EXPEDIENTE}' finaliza en {dias_restantes} días. Iniciar redacción."
                frecuencia = "semanal"
            elif 0 < dias_restantes <= 60:
                asunto = "📌 Finalización de contrato próxima"
                mensaje = f"El contrato '{c.EXPEDIENTE}' finaliza pronto. Quedan {dias_restantes} días."
                frecuencia = "semanal"

    if asunto and mensaje:
        print(f"✅ Generado correo para: {c.EXPEDIENTE} -> {asunto}")
        return {
            'asunto': asunto,
            'mensaje': mensaje,
            'frecuencia': frecuencia,
            'correo': "legal.fib.hlpr@salud.madrid.org"
        }

    print("⛔ No se envía correo.")
    return None




def restringir_seccion(seccion_requerida):
    def decorador(vista_func):
        @wraps(vista_func)
        def _wrapped_view(request, *args, **kwargs):
            usuario_id = request.session.get('usuario_id')
            if not usuario_id:
                return redirect('login')

            try:
                usuario = Usuario.objects.get(id=usuario_id)
                if str(seccion_requerida) not in usuario.secciones.split(','):
                    messages.error(request, "No tienes acceso a esta sección.")
                    return redirect('login')
            except Usuario.DoesNotExist:
                return redirect('login')

            return vista_func(request, *args, **kwargs)
        return _wrapped_view
    return decorador

# 🔧 Utilidades generales
def safe_str(val):
    return '' if pd.isna(val) else str(val).strip()

def limpiar_decimal(val):
    try:
        if pd.isna(val):
            return 0
        if isinstance(val, str):
            val = val.replace('.', '').replace(',', '.').replace('€', '').strip()
        return float(val)
    except Exception:
        return 0


def parse_fecha(valor):
    if pd.isna(valor) or not isinstance(valor, str):
        return None
    try:
        match = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', valor)
        if match:
            return datetime.strptime(match.group(1), "%d.%m.%Y").date()
        return datetime.strptime(valor.strip(), "%d.%m.%Y").date()
    except Exception:
        return None

def extraer_fecha(fecha_str):
    if not fecha_str:
        return None
    match = re.search(r"\b(\d{2}\.\d{2}\.\d{4})\b", fecha_str)
    if match:
        return parse_fecha(match.group(1))
    return parse_fecha(fecha_str)



# 📄 Convocatorias
@restringir_seccion('2')
def convocatorias_view(request):

    if request.method == 'POST' and 'excel_file' in request.FILES:
        df = pd.read_excel(request.FILES['excel_file'], sheet_name="2025", header=1)
        df.columns = df.columns.str.strip()
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df = df[~df.astype(str).apply(lambda row: row.str.contains("TOTAL", case=False, na=False)).any(axis=1)]
        df = df.dropna(how='all')
        df = df[df['CÓDIGO FUNDANET'].apply(lambda x: isinstance(x, str) and x.strip() != '')]

        for index, row in df.iterrows():
            codigo = safe_str(row.get('CÓDIGO FUNDANET'))
            if not codigo:
                continue
            try:
                Convocatoria.objects.update_or_create(
                    codigo_fundanet=codigo,
                    defaults={
                        'tipo': safe_str(row.get('TIPO')),
                        'nif_tercero': safe_str(row.get('NIF TERCERO')),
                        'ip': safe_str(row.get('IP')),
                        'nombre': safe_str(row.get('NOMBRE')),
                        'objeto': safe_str(row.get('OBJETO')),
                        'pais_ccaa': safe_str(row.get('PAIS/CCAA')),
                        'cantidad': limpiar_decimal(row.get('CANTIDAD')) if not pd.isna(row.get('CANTIDAD')) else 0,
                        'overhead': limpiar_decimal(row.get('CANTIDAD')) if not pd.isna(row.get('OVERHEAD')) else 0,
                        'ruta': safe_str(row.get('RUTA')),
                        'factura': safe_str(row.get('FACTURA (SI/NO)')),
                        'notas': safe_str(row.get('NOTAS')),
                        'estado_contabilidad': safe_str(row.get('ESTADO (CONTABILIDAD)')),
                        'contactos_fra': safe_str(row.get('CONTACTOS FRA'))
                    }

                )
            except Exception as e:
                print(f"❌ Error en fila {index}: {e}")
        return redirect('legal_convocatorias')

    form = ConvocatoriaForm()
    datos = Convocatoria.objects.all()
    resumen_por_tipo = defaultdict(lambda: {'cantidad': 0, 'total_cantidad': 0, 'total_overhead': 0})
    resumen_por_ip = defaultdict(lambda: {'cantidad': 0, 'total_cantidad': 0, 'total_overhead': 0})

    for c in datos:
        tipo = safe_str(c.tipo)
        resumen_por_tipo[tipo]['cantidad'] += 1
        resumen_por_tipo[tipo]['total_cantidad'] += c.cantidad or 0
        resumen_por_tipo[tipo]['total_overhead'] += c.overhead or 0

        for ip in re.split(r'\s*[\+\/]{1,2}\s*', c.ip or ''):
            ip = ip.strip()
            if ip:
                resumen_por_ip[ip]['cantidad'] += 1
                resumen_por_ip[ip]['total_cantidad'] += c.cantidad or 0
                resumen_por_ip[ip]['total_overhead'] += c.overhead or 0

    for resumen in [resumen_por_tipo, resumen_por_ip]:
        for data in resumen.values():
            data['total_combinado'] = data['total_cantidad'] + data['total_overhead']

    return render(request, 'legal/convocatorias.html', {
        'form': form,
        'datos': datos,
        'resumen_por_tipo': dict(resumen_por_tipo),
        'resumen_por_ip': dict(resumen_por_ip),
        'grafico_tipo_labels': list(resumen_por_tipo.keys()),
        'grafico_tipo_valores': [float(d['total_combinado']) for d in resumen_por_tipo.values()],
        'grafico_ip_labels': list(resumen_por_ip.keys()),
        'grafico_ip_valores': [float(d['total_combinado']) for d in resumen_por_ip.values()],
    })

# 📊 Concursos
def calcular_alerta_y_evento(c):
    estado = (c.ESTADO or '').strip().upper()
    if estado in ['CANCELADO', 'FINALIZADO', 'PARALIZADO', 'DESIERTO Y FINALIZADO', 'EN GESTIÓN']:
        return None, None

    titulo = f"{c.EXPEDIENTE}: {c.PROCEDIMIENTO}"
    prorroga = (c.PRORROGA or '').lower()
    fecha_objetivo = extraer_fecha(c.FECHA_FIN_PRORROGA if 'si' in prorroga else c.FECHA_FIN)

    if not fecha_objetivo:
        return None, None

    if '2 veces' in prorroga:
        alerta = fecha_objetivo - timedelta(days=425)
    elif 'si' in prorroga:
        alerta = fecha_objetivo - timedelta(days=90)
    else:
        alerta = fecha_objetivo - timedelta(days=120)

    return {
        'title': titulo,
        'start': fecha_objetivo.isoformat(),
        'color': 'blue',
    }, {
        'title': f"⚠️ Prórroga: {titulo}",
        'start': alerta.isoformat(),
        'color': 'red',
    }
@restringir_seccion('2')
def concursos_view(request):
    form = ConcursoForm()
    if request.method == 'POST' and 'excel_file' in request.FILES:
        df = pd.read_excel(request.FILES['excel_file'], header=2, usecols="B:L")
        for index, row in df.iterrows():
            try:
                Concurso.objects.update_or_create(
                    EXPEDIENTE=safe_str(row.get('N.º EXPED')),
                    defaults={
                        'PROCEDIMIENTO': safe_str(row.get('PROCEDIMIENTO')),
                        'TIPOLOGÍA': safe_str(row.get('TIPOLOGÍA')),
                        'FECHA_FIRMA': safe_str(row.get('FECHA FIRMA')),
                        'DURACIÓN_CONTRATO': safe_str(row.get('DURACIÓN CONTRATO')),
                        'FECHA_FIN': safe_str(row.get('FECHA FIN')),
                        'PRORROGA': safe_str(row.get('¿PRÓRROGAS?')),
                        'FECHA_FIN_PRORROGA': safe_str(row.get('FECHA FIN PRÓRROGA')),
                        'CANTIDAD_ECONOMICA_MAX': limpiar_decimal(row.get('CANTIDAD ECONÓMICA MÁX (IVA NO INCLUIDO)')),
                        'ESTADO': safe_str(row.get('ESTADO')),
                        'ESTADO_CONTINUACION': safe_str(row.get('ESTADO CONTINUACION'))
                    }
                )
            except Exception as e:
                print(f"❌ Error en fila {index}: {e}")
                print("➡️ Fila:", row.to_dict())
        return redirect('legal_concursos')
    elif request.method == 'POST' and 'editar_id' not in request.POST:
        # 🟪 Guardar concurso manual
        form = ConcursoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('legal_concursos')


    concursos = Concurso.objects.all()
    resumen_procedimiento = defaultdict(float)
    resumen_tipologia = defaultdict(float)
    resumen_anual = defaultdict(float)
    eventos = []

    for c in concursos:
        valor = float(c.CANTIDAD_ECONOMICA_MAX or 0)

        resumen_procedimiento[safe_str(c.PROCEDIMIENTO)] += valor
        resumen_tipologia[safe_str(c.TIPOLOGÍA)] += valor

        fecha = extraer_fecha(c.FECHA_FIRMA)
        if fecha:
            resumen_anual[fecha.year] += valor

        fin, aviso = calcular_alerta_y_evento(c)
        if fin:
            eventos.append(fin)
        if aviso:
            eventos.append(aviso)


    conteo_estado = defaultdict(int)

    for c in concursos:
        estado = safe_str(c.ESTADO)
        conteo_estado[estado] += 1


    hoy = date.today()
    mensajes_alerta = []

    for c in concursos:
        _, alerta = calcular_alerta_y_evento(c)
        if alerta and alerta['start'] <= hoy.isoformat():
            mensajes_alerta.append(alerta['title'])



    editar_id = request.GET.get('editar')

    if request.method == "POST" and 'editar_id' in request.POST:
        try:
            concurso = Concurso.objects.get(id=request.POST.get('editar_id'))

            campos_actualizables = [
                'procedimiento', 'tipologia', 'fecha_firma', 'duracion',
                'fecha_fin', 'prorroga', 'fecha_fin_prorroga',
                'cantidad', 'estado', 'estado_continuacion'
            ]

            for campo in campos_actualizables:
                valor = request.POST.get(campo)
                if valor is not None:
                    if campo == 'tipologia':
                        concurso.TIPOLOGÍA = valor.upper()
                    elif campo == 'estado':
                        concurso.ESTADO = valor.upper()
                    elif campo == 'cantidad':
                        concurso.CANTIDAD_ECONOMICA_MAX = valor or None
                    else:
                        setattr(concurso, campo.upper(), valor)

            concurso.save()
            print("DEBUG POST:", request.POST)
            return redirect('legal_concursos')

        except Concurso.DoesNotExist:
            pass



    return render(request, 'legal/concursos.html', {
        'concursos': concursos,
        'form': form,
        'editar_id': editar_id,
        'grafico_procedimiento': {
            'labels': list(resumen_procedimiento.keys()),
            'values': [round(v, 2) for v in resumen_procedimiento.values()]
        },
        'grafico_tipologia': {
            'labels': list(resumen_tipologia.keys()),
            'values': [round(v, 2) for v in resumen_tipologia.values()]
        },
        'grafico_anual': {
            'labels': sorted(resumen_anual.keys()),
            'values': [round(resumen_anual[year], 2) for year in sorted(resumen_anual.keys())]
        },
        'grafico_estado': {
            'labels': list(conteo_estado.keys()),
            'values': list(conteo_estado.values())
        },
        'eventos': eventos,
        'alertas_hoy': mensajes_alerta,
        'tipologias': ['MIXTO', 'SUMINISTRO', 'SERVICIO'],
        'estados': ['FINALIZADO', 'DESIERTO', 'ACTIVO', 'PARALIZADO', 'CANCELADO', 'DESIERTO Y FINALIZADO', 'EN GESTIÓN'],
        'prorrogas': ['NO', 'SI - 1 AÑO', 'SI - 18 MESES', 'SI - 2 AÑOS', 'SI - 1 AÑO - 2 VECES'],

    })

def calendario_concursos_view(request):
    concursos = Concurso.objects.all()
    eventos = []
    for c in concursos:
        fin, aviso = calcular_alerta_y_evento(c)
        if fin:
            eventos.append(fin)
        if aviso:
            eventos.append(aviso)
    return render(request, 'legal/calendario.html', {'eventos': eventos})

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse


def formato_europeo(valor):
    try:
        valor = float(valor)
        return "{:,.2f}".format(valor).replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return valor


def exportar_concursos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Concursos"

    # Encabezados
    headers = [
        "N.º EXPED", "PROCEDIMIENTO", "TIPOLOGÍA", "FECHA FIRMA",
        "DURACIÓN CONTRATO", "FECHA FIN", "¿PRÓRROGAS?", "FECHA FIN PRÓRROGA",
        "CANTIDAD ECONÓMICA MÁX", "ESTADO", "ESTADO CONTINUACION"
    ]
    ws.append(headers)

    # Estilo de encabezado
    header_fill = PatternFill(start_color="E9D8FD", end_color="E9D8FD", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Colores por estado
    estado_colores = {
        "FINALIZADO": "D1FAE5",         # verde claro
        "ACTIVO": "DBEAFE",             # azul claro
        "CANCELADO": "FECACA",          # rojo claro
        "DESIERTO": "F87171",           # rojo
        "PARALIZADO": "E5E7EB",         # gris claro
        "DESIERTO Y FINALIZADO": "FBCFE8",  # rosa claro
        "EN GESTION": "FEF3C7",         # amarillo claro
    }

    concursos = Concurso.objects.all()
    for row_index, c in enumerate(concursos, start=2):
        fila = [
            c.EXPEDIENTE, c.PROCEDIMIENTO, c.TIPOLOGÍA, c.FECHA_FIRMA,
            c.DURACIÓN_CONTRATO, c.FECHA_FIN, c.PRORROGA, c.FECHA_FIN_PRORROGA,
            formato_europeo(c.CANTIDAD_ECONOMICA_MAX),
            c.ESTADO, c.ESTADO_CONTINUACION
        ]
        ws.append(fila)

        estado = (c.ESTADO or "").strip().upper()
        fill_color = estado_colores.get(estado, None)

        for col_index, cell in enumerate(ws[row_index], start=1):
            cell.alignment = center_align
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Ajuste automático de ancho (opcional, puede fallar con mucho texto)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 30)

    # Exportar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Concursos.xlsx"'
    wb.save(response)
    return response


from django.shortcuts import get_object_or_404

@restringir_seccion('2')
def borrar_concurso(request, pk):
    concurso = get_object_or_404(Concurso, pk=pk)
    concurso.delete()
    return redirect('legal_concursos')


@restringir_seccion('2')
def borrar_convocatoria(request, pk):
    convocatoria = get_object_or_404(Convocatoria, pk=pk)
    convocatoria.delete()
    return redirect('legal_convocatorias')

@restringir_seccion('2')
def editar_convocatoria(request, pk):
    convocatoria = get_object_or_404(Convocatoria, pk=pk)
    if request.method == "POST":
        form = ConvocatoriaForm(request.POST, instance=convocatoria)
        if form.is_valid():
            form.save()
            return redirect('legal_convocatorias')
    else:
        form = ConvocatoriaForm(instance=convocatoria)

    return render(request, 'legal/editar_convocatoria.html', {
        'form': form,
        'convocatoria': convocatoria
    })